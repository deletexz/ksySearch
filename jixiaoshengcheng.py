import datetime
import re
import sys
import openpyxl
# import time
# from datetime import datetime
import xlwt

# 定义类
class News:
    def __init__(self, id,name,cehua,wenan,paishe,baozhuang,amount,cehuSalary,wenanSalary,paisheSalary,baozhuangSalary,
                 xz1,xzSa1,xz2,xzSa2,xz3,xzSa3,xz4,xzSa4):

        self.id = id
        self.name = name
        self.cehua = cehua
        self.wenan = wenan
        self.paishe = paishe
        self.baozhuang = baozhuang
        self.amount = amount

        self.cehuSalary = cehuSalary
        self.wenanSalary = wenanSalary
        self.paisheSalary = paisheSalary
        self.baozhuangSalary = baozhuangSalary

        self.xz1 = xz1
        self.xzSa1 = xzSa1
        self.xz2 = xz2
        self.xzSa2 = xzSa2
        self.xz3 = xz3
        self.xzSa3 = xzSa3
        self.xz4 = xz4
        self.xzSa4 = xzSa4


# 储存对象
Arraylistss = []

NameSalary = {}
NameSalarylist = {}


# 处理筛选数据，存入对象
def Data_processing():

    # 创建一个工作簿
    # wb = openpyxl.Workbook()
    # 创建一个test_case的sheet表单
    # wb.create_sheet('test_case')
    # 保存为一个xlsx格式的文件
    # wb.save('cases.xlsx')
    # 读取excel中的数据
    # 第一步：打开工作簿
    wb = openpyxl.load_workbook('E:\code_project\Python\ksySearch\cases.xlsx')
    # wb = openpyxl.load_workbook('cases.xlsx')
    # 第二步：选取表单
    sh = wb['Sheet1']
    # 第三步：读取数据
    # 参数 row:行  column：列
    # id = sh.cell(row = 3,column = 1)   # 读取第一行，第一列的数据
    # name = sh.cell(row = 3,column = 3)   # 读取第一行，第一列的数据
    # cehua = sh.cell(row = 3,column = 4)
    # wenan = sh.cell(row = 3,column = 5)
    # paishe = sh.cell(row = 3,column = 6)
    # baozhuang = sh.cell(row = 3,column = 7)
    # amount = sh.cell(row = 3,column = 16)


    # print(list(sh.rows)[1:])     # 按行读取数据，去掉第一行的表头信息数据
    for lines in list(sh.rows)[2:]:

        id = lines[0].value
        if(id == None):
            continue
        if(id == '序号'):
            continue
        if(id == ''):
            continue
        name = lines[2].value
        cehuatext = lines[3].value
        wenantext = lines[4].value
        paishetext = lines[5].value
        baozhuangtext = lines[6].value
        amount = lines[14].value

        xz1text = lines[15].value
        xz2text = lines[16].value
        xz3text = lines[17].value
        xz4text = lines[18].value

        # 关闭工作薄
        wb.close()
        # amount = lines[8].value
        if(amount ==None):
            continue


        cehualist = cehuatext.split('（', 1 ) if cehuatext != None else ''# 以'（'为分隔符，分隔成两个
        wenanlist = wenantext.split('（', 1 ) if wenantext != None else ''
        paishelist = paishetext.split('（', 1 ) if paishetext != None else ''
        baozhuanglist = baozhuangtext.split('（', 1 ) if baozhuangtext != None else ''


        #提取姓名
        cehua = re.sub('[^\u4e00-\u9fa5]+','',cehualist[0] if len(cehualist)>1 else '')
        wenan = re.sub('[^\u4e00-\u9fa5]+','',wenanlist[0] if len(wenanlist)>1 else '')
        paishe = re.sub('[^\u4e00-\u9fa5]+','',paishelist[0] if len(paishelist)>1 else '')
        baozhuang = re.sub('[^\u4e00-\u9fa5]+','',baozhuanglist[0] if len(baozhuanglist)>1 else '')
        # cehua = cehualist[0] if len(cehualist)>1 else ''
        # wenan = wenanlist[0] if len(wenanlist)>1 else ''
        # paishe = paishelist[0] if len(paishelist)>1 else ''
        # baozhuang = baozhuanglist[0] if len(baozhuanglist)>1 else ''

        #提取金额
        cehuSalary = re.findall(r'\d+', cehualist[1] if len(cehualist)>1 else '0')
        wenanSalary = re.findall(r'\d+', wenanlist[1] if len(wenanlist)>1 else '0')
        paisheSalary = re.findall(r'\d+', paishelist[1] if len(paishelist)>1 else'0')
        baozhuangSalary= re.findall(r'\d+', baozhuanglist[1] if len(baozhuanglist)>1 else '0')

        cehuSalary = float(cehuSalary[0]) /100 * float(amount)
        wenanSalary = float(wenanSalary[0]) /100 * float(amount)
        paisheSalary = float(paisheSalary[0]) /100 * float(amount)
        baozhuangSalary= float(baozhuangSalary[0]) /100 * float(amount)

        xz = {}
        for insz in range(xz1text,xz2text,xz3text,xz4text):

            xzl = re.sub('[^\u4e00-\u9fa5]+','',insz[0] if len(cehualist)>1 else '')
            xzlSa = re.findall(r'\d+', insz[1] if len(cehualist)>1 else '0')
            xzlSalary = float(xzlSa[0]) /100 * float(amount)

            if(xzl != '' and xzlSalary != 0):
                # xzdi = {}
                xz[xzl] = xzlSalary



        xzpes = xz.keys()

        if(len(xzpes) == 0):
            xz1 = ''
            xzSa1 = 0
            xz2 = ''
            xzSa2 = 0
            xz3 = ''
            xzSa3 = 0
            xz4 = ''
            xzSa4 = 0
        elif(len(xzpes) == 1):
            xz1 = xzpes[0]
            xzSa1 = xz[xzpes[0]]

            xz2 = ''
            xzSa2 = 0
            xz3 = ''
            xzSa3 = 0
            xz4 = ''
            xzSa4 = 0
        elif(len(xzpes) == 2):
            xz1 = xzpes[0]
            xzSa1 = xz[xzpes[0]]
            xz2 = xzpes[1]
            xzSa2 = xz[xzpes[1]]

            xz3 = ''
            xzSa3 = 0
            xz4 = ''
            xzSa4 = 0
        elif(len(xzpes) == 3):
            xz1 = xzpes[0]
            xzSa1 = xz[xzpes[0]]
            xz2 = xzpes[1]
            xzSa2 = xz[xzpes[1]]
            xz3 = xzpes[2]
            xzSa3 = xz[xzpes[2]]

            xz4 = ''
            xzSa4 = 0

        else:
            xz1 = xzpes[0]
            xzSa1 = xz[xzpes[0]]
            xz2 = xzpes[1]
            xzSa2 = xz[xzpes[1]]
            xz3 = xzpes[2]
            xzSa3 = xz[xzpes[2]]
            xz4 = xzpes[3]
            xzSa4 = xz[xzpes[3]]

        print(id)
        news = News(id,name,cehua,wenan,paishe,baozhuang,amount,cehuSalary,wenanSalary,paisheSalary,baozhuangSalary,
                    xz1,xzSa1,xz2,xzSa2,xz3,xzSa3,xz4,xzSa4)
        print(news)
        print(Arraylistss)
        Arraylistss.append(news)


        # print(Arraylistss[0])
        # print("")




    # print(Arraylistss.values())

    # print(id.value)
    # print(cehua.value)
    # print()
    # 按行读取数据 list(sh.rows)
    # print(list(sh.rows)[1:])     # 按行读取数据，去掉第一行的表头信息数据
    # for cases in list(sh.rows)[1:]:
    #     case_id =  cases[0].value
    #     case_excepted = cases[1].value
    #     case_data = cases[2].value
    #     print(case_excepted,case_data)
    # # 关闭工作薄
    # wb.close()


# 汇总
# def Salary2(s,k):
#     for pers in Arraylistss:
#         # id,name,cehua,wenan,paishe,baozhuang,amount,cehuSalary,wenanSalary,paisheSalary,baozhuangSalary
#         for per in range(pers.cehua,pers.wenan,pers.paishe,pers.baozhuang):
#             if(NameSalary.get(per) == None):
#                 NameSalary[per] = pers.cehuSalary
#                 list1 = [per.cehua]
#                 list1.append(per.cehuSalary)
#                 NameSalarylist[per.cehua] = list1
#             else:
#                 NameSalary[per.cehua] = NameSalary[per.cehua]+per.cehuSalary   #合计
#                 if(per.cehuSalary == 0):
#                     continue
#                 NameSalarylist[per.cehua].append(per.cehuSalary)  #明细
#
#     print(NameSalary.values())
#     print(NameSalarylist.values())
#     # print(Arraylistss[0])
#     print("")
#     pass

# 汇总
def Salary(s,k):
    for per in Arraylistss:
        if(NameSalary.get(per.cehua) == None):
            NameSalary[per.cehua] = per.cehuSalary
            list1 = [per.cehua]
            list1.append({per.id:per.cehuSalary})
            NameSalarylist[per.cehua] = list1
        else:
            NameSalary[per.cehua] = NameSalary[per.cehua]+per.cehuSalary             #合计
            if(per.cehuSalary != 0):
                NameSalarylist[per.cehua].append({per.id:per.cehuSalary})  #明细


        if(NameSalary.get(per.wenan) == None):
            NameSalary[per.wenan] = per.wenanSalary
            list1 = [per.wenan]
            list1.append({per.id:per.wenanSalary})
            NameSalarylist[per.wenan] = list1

        else:
            NameSalary[per.wenan] = NameSalary[per.wenan]+per.wenanSalary
            if(per.wenanSalary != 0):
                NameSalarylist[per.wenan].append({per.id:per.wenanSalary})

        if(NameSalary.get(per.paishe) == None):
            NameSalary[per.paishe] = per.paisheSalary
            list1 = [per.paishe]
            list1.append({per.id:per.paisheSalary})
            NameSalarylist[per.paishe] = list1

        else:
            NameSalary[per.paishe] = NameSalary[per.paishe]+per.paisheSalary
            if(per.paisheSalary != 0):
                NameSalarylist[per.paishe].append({per.id:per.paisheSalary})


        if(NameSalary.get(per.baozhuang) == None):
            NameSalary[per.baozhuang] = per.baozhuangSalary
            list1 = [per.baozhuang]
            list1.append({per.id:per.baozhuangSalary})
            NameSalarylist[per.baozhuang] = list1

        else:
            NameSalary[per.baozhuang] = NameSalary[per.baozhuang]+per.baozhuangSalary
            if(per.baozhuangSalary != 0):
                NameSalarylist[per.baozhuang].append({per.id:per.baozhuangSalary})



    print(NameSalary.values())
    print(NameSalarylist.values())
    # print(Arraylistss[0])
    print("")
    pass


def Sumlist():

    namelist = NameSalary.keys()
    for nm in namelist:
        NameSalarylist[nm].insert(0,NameSalary[nm])

    # print(NameSalarylist)


    new_wk=xlwt.Workbook()  #创建工作簿
    new_sheet=new_wk.add_sheet('sheetname')    #创建名为sheetname的工作表
    new2_sheet2=new_wk.add_sheet('sheetname2')
    new3_sheet3=new_wk.add_sheet('sheetname3')
    i=0
    for nm in NameSalarylist:

        if(nm==''):
            continue
        new_sheet.write(i,0,nm)
        new2_sheet2.write(i,0,nm)
        new3_sheet3.write(i,0,nm)
        j=1
        for NSl in NameSalarylist[nm]:
            if(j>2):
                new_sheet.write(i,j,list(NSl.values())[0])
                new2_sheet2.write(i,j,list(NSl)[0])
                new3_sheet3.write(i,j,str(NSl))
            else:
                new_sheet.write(i,j,NSl)  #在i行，j列写入内容：content，i，j从0开始
                new2_sheet2.write(i,j,NSl)
                new3_sheet3.write(i,j,NSl)
            j=j+1

        i=i+1

    new_wk.save('结果.xls')     #使用xlwt写入操作后，需要保存


    # # 创建一个工作薄：
    # wb = openpyxl.Workbook()
    # # 新增一个sheet表单：
    # wb.create_sheet('sheet')
    # # 保存case.xlsx文件：
    # wb.save('结果.xlsx')




#
FaGaolist = []
# 发稿等级统计
# A 5000 3000
# B 1000
# C 300
# D 200
# E 150
# F 100
A = []
B = []
C = []
D = []
E = []
F = []
def FaGao():
    for per in Arraylistss:
        if(per.amount == 5000 or per.amount == 3000):
            litfg(A,per,"A")
        elif(per.amount ==1000):
            litfg(B,per,"B")
        elif(per.amount ==300):
            litfg(C,per,"C")
        elif(per.amount ==200):
            litfg(D,per,"D")
        elif(per.amount ==150):
            litfg(E,per,"E")
        elif(per.amount ==100):
            litfg(F,per,"F")
        else:
            continue

    # A = [{name:**,id:**},{name:**,id:**}]
    bijiao = {}
    # for i in range(0,len(A)):

    # count= []
    counts = {}
    # Namedic = {}

    listABCEDF = [A,B,C,D,E,F]
    # 循环[A,B,C,D,E,F]
    for perABCEDF in listABCEDF:

        # 循环 B
        for i in perABCEDF:
            bijiaoName=i["name"]
            bijiaoId=i["id"]
            bigrade=i["Grade"]

            # {name:{{"A":[2,3,4]},{"B":[2,3,4]}}}
            Namedic = {bijiaoName:{}}

            # cot={bijiaoName:[{bigrade:[bijiaoId]}]}

            # {"A":[2,3,4]}
            Dic = {bigrade:[bijiaoId]}
            for ii in perABCEDF:
                if(bijiaoName == ii["name"]):
                    Dic[bigrade].append((ii["id"]))

                    # cot[bijiaoName][bigrade].append(ii["id"])
            # cot[bijiaoName] = set(cot[bijiaoName])
            Dic[bigrade] = list(set(Dic[bigrade]))
            # {name:[{"A":[2,3,4]},{"B":[2,3,4]}]}
            # Namedic[bijiaoName]

            Namedic[bijiaoName].update(Dic)

            # cot[bijiaoName][cotindex][bigrade] = list(set(cot[bijiaoName][bigrade]))

            if(counts.get(bijiaoName)!=None):
                counts[bijiaoName].update(Dic)
            else:
                counts.update(Namedic)
                pass

            # count.append(Namedic)
            # cot.clear()

    finalcout = {}
    for cname in counts.keys():
        ccna = {cname: {}}
        for cgrade in counts[cname].keys():
            clen = len(counts[cname][cgrade])
            ccgr = {cgrade:clen}

            ccna[cname].update(ccgr)

        finalcout.update(ccna)


    new_wk=xlwt.Workbook()  #创建工作簿
    new_sheet=new_wk.add_sheet('sheetname')    #创建名为sheetname的工作表

    #在i行，j列写入内容：content，i，j从0开始
    # new_sheet.write(i,j,NSl)
    j=1
    for perabc in ["A","B","C","D","E","F"]:
        new_sheet.write(0,j,perabc)
        j=j+1
    j=1
    for perna in finalcout.keys():

        new_sheet.write(j,0,perna)

        sum = 0
        for per in finalcout[perna]:
            if(per=="A"):
                text = "A类稿件"+ str(finalcout[perna][per]) +"件;"
                new_sheet.write(j,1,text)
                sum =sum + finalcout[perna][per]
            elif(per=="B"):
                text = "B类稿件"+ str(finalcout[perna][per]) +"件;"
                new_sheet.write(j,2,text)
                sum =sum + finalcout[perna][per]
            elif(per=="C"):
                text = "C类稿件"+ str(finalcout[perna][per]) +"件;"
                new_sheet.write(j,3,text)
                sum =sum + finalcout[perna][per]
            elif(per=="D"):
                text = "D类稿件"+ str(finalcout[perna][per]) +"件;"
                new_sheet.write(j,4,text)
                sum =sum + finalcout[perna][per]
            elif(per=="E"):
                text = "E类稿件"+ str(finalcout[perna][per]) +"件;"
                new_sheet.write(j,5,text)
                sum =sum + finalcout[perna][per]
            elif(per=="F"):
                text = "F类稿件"+ str(finalcout[perna][per]) +"件;"
                new_sheet.write(j,6,text)
                sum =sum + finalcout[perna][per]
            else:
                pass

        new_sheet.write(j,8,sum)

        j=j+1

    new_wk.save('等级.xls')     #使用xlwt写入操作后，需要保存
    # count.__str__()
    print("")



# 抽象
def litfg(listfg,per,djGrade):

    # self.cehua = cehua
    # self.wenan = wenan
    # self.paishe = paishe
    # self.baozhuang = baozhuang
    if(per.cehua!=""):
        listfg.append({"id":per.id,"name":per.cehua,"Grade":djGrade})
    if(per.wenan!=""):
        listfg.append({"id":per.id,"name":per.wenan,"Grade":djGrade})
    if(per.paishe!=""):
        listfg.append({"id":per.id,"name":per.paishe,"Grade":djGrade})
    if(per.baozhuang!=""):
        listfg.append({"id":per.id,"name":per.baozhuang,"Grade":djGrade})
        


if __name__ == "__main__":
    # starttime = datetime.datetime.now().days
    # endtime = datetime.datetime.now()
    # print (endtime - starttime).seconds

    # time.strftime("%Y-%m-%d %H:%M:%S", time.localtime())
    # starttime = time.strftime("%Y-%m-%d", time.localtime()).date()
    # starttime = datetime.datetime.now()

    starttime = datetime.date.today()
    endtime = datetime.datetime.strptime('2022-01-01', '%Y-%m-%d').date()
    if(endtime < starttime):
        # print("程序结束")
        sys.exit()

    # print("程序开始")
    # 提取数据
    Data_processing()
    # 处理数据
    Salary(0,1)
    Sumlist()

    # 等级统计，发稿等级
    FaGao()



    print("")
    print("")











